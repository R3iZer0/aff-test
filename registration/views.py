from aiohttp import request
from django.urls import path
from django.shortcuts import redirect, render
from .forms import LeadForm
from .models import Lead
from django.http import HttpRequest, HttpResponse, request
from django.http import JsonResponse
from openpyxl import Workbook
import os
import datetime
import pytz
from googleapiclient.discovery import build
from google.oauth2 import service_account
from django.shortcuts import render
from .models import Lead

def home(request):
    return render(request, 'home.html', {})




from django.shortcuts import render
from .forms import LeadForm

def lead_register(request):
    if request.method == 'POST':
        form = LeadForm(request.POST)
        if form.is_valid():
            # Save the lead to the database
            form.save()
            # Redirect to a "thank you" page or some other page
            return redirect('registration')
    else:
        form = LeadForm()
    return render(request, 'registration.html', {'form': form})


from django.shortcuts import render

def leads_data_view(request):
    leads = Lead.objects.all()  # Replace 'Lead' with your actual model name or queryset
    context = {
        'leads': leads
    }
    return render(request, 'leads_data.html', context)




def lead_list(request):
    leads = Lead.objects.all()
    return render(request, 'leads_view.html', {'leads': leads})


def download_table(request):
    leads = Lead.objects.all()  # Replace 'Lead' with your actual model name or queryset
    
    # Create a new workbook and get the active sheet
    workbook = Workbook()
    sheet = workbook.active
    
    # Write the table headers
    headers = ['Id', 'Name', 'Last Name', 'Email', 'Phone', 'Country', 'Experience', 'Registered at']
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num, value=header)
    
    # Write the table data
    for row_num, lead in enumerate(leads, 2):
        sheet.cell(row=row_num, column=1, value=lead.id)
        sheet.cell(row=row_num, column=2, value=lead.name)
        sheet.cell(row=row_num, column=3, value=lead.last_name)
        sheet.cell(row=row_num, column=4, value=lead.email)
        sheet.cell(row=row_num, column=5, value=lead.phone)
        sheet.cell(row=row_num, column=6, value=lead.country)
        sheet.cell(row=row_num, column=7, value=lead.experience)
        created_at_naive = lead.created_at.replace(tzinfo=None)
        sheet.cell(row=row_num, column=8, value=created_at_naive)    
    # Create the response object with appropriate headers
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=leads.xlsx'
    
    # Save the workbook to the response object
    workbook.save(response)
    
    return response





def save_to_google_sheet(request: HttpRequest):
    # Path to the service account key JSON file
    key_file_path = 'registration/static/css/i-informatics-387814-d693dbfa6c95.json'

    # ID of the Google Sheet to be updated
    sheet_id = '13PQeiimRhhPBJ86eUSPMUpNAFQh0-zB0Lkrq1CzsPZo'

    # Connect to Google Sheets API using the service account key
    credentials = service_account.Credentials.from_service_account_file(key_file_path)
    service = build('sheets', 'v4', credentials=credentials)

    # Get the current date and time in UTC
    current_datetime = datetime.datetime.now(pytz.utc).strftime('%Y-%m-%d %H:%M:%S')

    # Retrieve the table data from the database
    leads = Lead.objects.all()

    # Prepare the data to be written to the Google Sheet
    data = []
    for lead in leads:
        data.append([lead.name, lead.last_name, lead.email, lead.phone, lead.country, lead.get_experience_display(), current_datetime])

    # Define the range where the data should be written (e.g., Sheet1!A1:G)
    range_name = 'Sheet1!A1:G'

    # Construct the request to update the Google Sheet
    value_range_body = {
        'values': data
    }
    request = service.spreadsheets().values().update(
        spreadsheetId=sheet_id,
        range=range_name,
        valueInputOption='RAW',
        body=value_range_body
    )
    response = request.execute()

    # Render a response to the user
    return render(request, 'leads_view.html')